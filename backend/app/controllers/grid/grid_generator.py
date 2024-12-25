"""
This file is responsible for generating a grid-based Excel sheet that displays a student's class schedule.

Overview:
- The script uses the OpenPyXL library to modify a pre-existing Excel template file ("Timetable template.xlsx").
- It parses the student's availability (fetched using their unique `studentId`) and marks the schedule grid to indicate their availability using color-coded cells.
- The script supports clearing the grid (resetting it to a blank template) and filling it based on the student's availability.

Key Functions:
1. `fill_in_day`: Marks the cells in the grid for a specific day based on the student's availability.
2. `is_available`: Checks if a specific time falls within the student's availability ranges.
3. `fill_in_cell`: Colors a specific cell in the worksheet.
4. `clear_row`: Resets all cells in a specific row (day) to a blank state.
5. `clear_grid`: Clears the entire grid, resetting all rows and columns.
6. `fill_in_schedule`: Updates the grid with a student's availability fetched using their `studentId`.

Dependencies:
- OpenPyXL: For manipulating Excel files.
- Custom parser (from`controllers.grid_generator.helperclasses.availability_parser.py`): For fetching and parsing a student's availability.

Inputs:
- `studentId`: A unique identifier for the student, entered manually via the console.
- Excel template: The script expects the template file `Timetable template.xlsx` to exist in the same directory.

Outputs:
- A modified Excel file with the grid updated to reflect the student's schedule.

Usage:
- Run the script directly, and it will prompt the user to enter a `studentId`.
- The updated Excel file will be saved with the same name, overwriting the template.

"""
#------------------------------------------------------------------------ Imports ------------------------------------------------------------------------#
import sys
import os
from pathlib import Path
import datetime
from openpyxl.styles import PatternFill
from openpyxl import load_workbook

# Get the absolute path of the current file and resolve any symlinks
# Then navigate up 3 levels to reach app directory:
# grid_generator.py -> grid -> controllers -> app
root_dir = Path(__file__).resolve().parents[2]
sys.path.append(str(root_dir))

# Now we can import our local modules
from controllers.grid.helper_classes.availability_parser import parse_availability

#------------------------------------------------------------------------ Constants ------------------------------------------------------------------------#
FINAL_HOUR_IN_GRID = 22
NUM_INTERVALS_PER_HOUR = 6
GRID_FILE_NAME = Path(__file__).parent / "Timetable template.xlsx"
GRID_FILL_COLOR = "ffa07a"


def fill_in_day(ws, dayId, availableRanges, color): 
    """
    Fills in the day/row for the students' class schedule based on the student's availability

    Parameters:
        ws: openpyxl.Worksheet
            The Excel worksheet object to update.
        dayId: int
            The integer id representing the day of week
        availableRanges: list
            The list of start/end times for when the student is available (i.e. List of classes for the day)

    Returns:
        None

    Side Effects:
        - Modifies the provided Excel workbook in place.
    """
    # Sunday (ID = 1) starts at row 3 (1 + 2) = 3
    rowNum = dayId + 2
    colNum = 2

    for hour in range(6, FINAL_HOUR_IN_GRID):
        minute = 0
        while(minute < 60):
            currentTime = datetime.time(hour=hour, minute=minute)

            if (not is_available(currentTime, availableRanges)):
                fill_in_cell(ws, rowNum, colNum, color)

            minute += 5
            colNum += 1


def is_available(currentTime, availableRanges):
    """
     Determines whether a given time falls within any of the specified availability ranges.

     This function checks if the provided `currentTime` is within the start and end times
     of any time range in the `availableRanges` list.

     Parameters:
         currentTime (datetime.time):
             The time to check for availability.
         availableRanges (list of dict):
             A list of dictionaries, where each dictionary represents a time range and contains:
                 - `'start_time'`: A `datetime.time` object marking the start of the range.
                 - `'end_time'`: A `datetime.time` object marking the end of the range.

             Example:
             [
                 {'start_time': datetime.time(9, 0), 'end_time': datetime.time(12, 0)},
                 {'start_time': datetime.time(14, 0), 'end_time': datetime.time(16, 0)}
             ]

     Returns:
         bool:
             - `True` if `currentTime` falls within any of the time ranges in `availableRanges`.
             - `False` if `currentTime` is outside all the ranges.
    """

    for rng in availableRanges:
        if rng["end_time"] > currentTime >= rng["start_time"]:
            return True

    return False


def fill_in_cell(ws, row, col, color):
    """
    Fills the cell (specified by row and col) in the openpyxl worksheet with the specified color.

    Parameters:
        ws: openpyxl.Worksheet
             The Excel worksheet object to update.
        row: int
            The row of the cell to be filled
        col: int
            the column of the cell to be filled
        color:
            the color of the cell to be filled specified by a 6-digit hex code

    Returns:
        None

    Side Effects:
        - Modifies the excel file directly

     """
    # Create a PatternFill object with the given color (in hexadecimal format)
    fill_color = PatternFill(start_color=color, end_color=color, fill_type="solid")

    # Use the row and column to get the cell and apply the fill color
    ws.cell(row=row, column=col).fill = fill_color


def clear_row(ws, dayId):
    """
    Resets the grid within the scope of one row

    Parameters:
        ws: openpyxl.Worksheet
            The Excel worksheet object to update.
        dayId: int
            The integer id representing the day of week

    Returns:
        None

    Side Effects:
        - Modifies the provided Excel workbook in place.

    Returns none

    """

    rowNum = dayId + 2
    colNum = 2

    for hour in range(6, FINAL_HOUR_IN_GRID):
        minute = 0
        while(minute < 60):
            # if we are on Monday, Wednesday, or Friday
            if dayId % 2 == 0:
                fill_in_cell(ws, rowNum, colNum, "FFFFFF")
            elif dayId == 1 or dayId == 7:
                fill_in_cell(ws, rowNum, colNum, "bababa")
            else:
                fill_in_cell(ws, rowNum, colNum, "e0e0e0")

            colNum += 1
            minute += 5


def clear_grid(ws):
    """
    Resets the grid to a blank template

    Parameters:
        ws: openpyxl.Worksheet
            The Excel worksheet object to update.

    Returns:
        None

    Side Effects:
        - Modifies the provided Excel workbook in place.
    """
    for i in range(7):
        dayId = i + 1
        clear_row(ws, dayId)


def fill_in_schedule(ws, studentId, color):
    """
    Updates a schedule in an Excel workbook based on a student's availability.

    Parameters:
        ws: openpyxl.Worksheet
            The Excel worksheet object to update.
        studentId: str
            Unique identifier for the student
        color: str or openpyxl.styles.Color
            The color to use when marking availability in the workbook.

    Returns:
        None

    Side Effects:
        - Modifies the provided Excel workbook in place.

    Exceptions:
        - If an error occurs during parsing availability or writing to the workbook,
          the function may raise an app aropriate exception.
    """
    try:
        avail = parse_availability(studentId)
        if avail:
            for day in avail:
                print(day)
                fill_in_day(ws, day["DayId"], day["DayRanges"], color)
        else:
            print("NO AVAILABILITY PARSED")
    except Exception as e:
        print("ERROR OCCURRED WHILE FILLING IN SCHEDULE: ", e)


def main():
    """
    Main file for running and debugging the grid generator on its own
    """

    wb = load_workbook(GRID_FILE_NAME)
    ws = wb.active
    clear_grid(ws)

    studentId = input("Enter the Student ID Number: ")
    fill_in_schedule(ws, studentId, GRID_FILL_COLOR)

    wb.save(GRID_FILE_NAME)


if __name__ == "__main__":
    main()

