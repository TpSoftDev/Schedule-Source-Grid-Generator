# This file is responsible for generating time grids for a new hires class schedule
# The programatic equivilant to manually highlighting a students grid on paper
# Generates a new time table in the project directory

#We nee
import datetime

from openpyxl import load_workbook
from datetime import time

from openpyxl.styles import PatternFill
import os
import sys
from controllers.grid_generator.helper_classes.availability_parser import parse_availability

from openpyxl import load_workbook
from controllers.grid_generator.helper_classes.availability_parser import parse_availability

FINAL_HOUR_IN_GRID = 22
NUM_INTERVALS_PER_HOUR = 6


def fill_in_day(wb, dayId, availableRanges, color):
    ws = wb.active
    # Sunday (ID = 1) starts at row 3 (1 + 2) = 3
    rowNum = dayId + 2
    colNum = 2

    for hour in range(6, FINAL_HOUR_IN_GRID):
        minute = 0
        while(minute < 60):
            currentTime = datetime.time(hour=hour, minute=minute)

            if (not is_available(currentTime, availableRanges)):
                fill_in_cell(ws, rowNum, colNum, color)

            minute += 5
            colNum += 1

    wb.save("Timetable template.xlsx")


#TODO Document
def is_available(currentTime, availableRanges):
    for range in availableRanges:
        if currentTime < range["end_time"] and currentTime >= range["start_time"]:
            return True

    return False


#TODO document
def fill_in_cell(ws, row, col, color):
    # Create a PatternFill object with the given color (in hexadecimal format)
    fill_color = PatternFill(start_color=color, end_color=color, fill_type="solid")

    # Use the row and column to get the cell and apply the fill color
    ws.cell(row=row, column=col).fill = fill_color



def clear_row(wb, dayId):
    ws = wb.active
    rowNum = dayId + 2
    colNum = 2

    for hour in range(6, FINAL_HOUR_IN_GRID):
        minute = 0
        while(minute < 60):
            # if we are on Monday, Wednesday, or Friday
            if dayId % 2 == 0:
                fill_in_cell(ws, rowNum, colNum, "FFFFFF")
            elif dayId == 1 or dayId == 7:
                fill_in_cell(ws, rowNum, colNum, "bababa")
            else:
                fill_in_cell(ws, rowNum, colNum, "e0e0e0")

            colNum += 1
            minute += 5

    wb.save("Timetable template.xlsx")


def clearGrid(wb):
    for i in range(7):
        dayId = i + 1
        clear_row(wb, dayId)

    wb.save("Timetable template.xlsx")


def fillInSchedule(wb, studentId, color):
    avail = parse_availability(studentId)
    pass
    if avail:
        for day in avail:
            print(day)
            fill_in_day(wb, day["DayId"], day["DayRanges"], color)
    else :
        print("NO AVAILABILITY")

    wb.save("Timetable template.xlsx")



def main():
    wb = load_workbook("Timetable template.xlsx")

    clearGrid(wb)
    color = "ffa07a"

    studentId = input("Enter the Student ID Number: ")
    fillInSchedule(wb, studentId, color)

    wb.save("Timetable template.xlsx")




# Run the main program when gridGenerator is executed
if __name__ == "__main__":
    main()