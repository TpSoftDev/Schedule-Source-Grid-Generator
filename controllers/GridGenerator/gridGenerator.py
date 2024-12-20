# This file is responsible for generating time grids for a new hires class schedule
# The programatic equivilant to manually highlighting a students grid on paper
# Generates a new time table in the project directory
#TODO: This Python file was made for the old class schedule data, Needs Refactoring
#We nee

from openpyxl import load_workbook
from datetime import time

from openpyxl.styles import PatternFill
import os
import sys

# Get the current directory of the script
current_dir = os.path.dirname(os.path.abspath(__file__))
external_directory = os.path.join(current_dir, "..")
sys.path.append(external_directory)

from api_calls.workday_api.workday_api import getStudentSchedule
from utils.helperFunctions import convert_to_time


# Iterate through each day to fill in the grid
# Param: "data" - the student employee's class schedule retrieved from a Workday API Call
# No return value
def populateGrid(data, ws):
    for course in data:
        fillInDay(course, "U", 3, ws)
        fillInDay(course, "M", 4, ws)
        fillInDay(course, "T", 5, ws)
        fillInDay(course, "W", 6, ws)
        fillInDay(course, "R", 7, ws)
        fillInDay(course, "F", 8, ws)
        fillInDay(course, "S", 9, ws)


# Iterates through each time slot in the "day" row of the grid and determines whether or not to highlight that slot
# Param : "course" - the class section object we want to block out time for
#        "day" - specifies which day we are looking at to compare the class schedule to
#        "rowNum" - determines which row in the excel sheet to highlight
# No return value
def fillInDay(course, day, rowNum, ws):
    max_col = ws.max_column
    hour = 6
    for col_num_outer in range(2, max_col + 1, 12):  # hour loop:
        minute = 0
        for col_num_inner in range(col_num_outer, col_num_outer + 12):
            # print("TIME SLOT: " + str(timeSlot) + "  Hour: " + str(hour) + "  Minute: " + str(minute)+ "\n")
            currentTime = time(hour, minute)
            if not (isAvailable(course, day, currentTime)):
                cell = ws.cell(row=rowNum, column=col_num_inner)
                fillColor = PatternFill(
                    start_color="98FF98", end_color="98FF98", fill_type="solid"
                )
                cell.fill = fillColor

            minute += 5

        hour += 1


# Observes one time slot on the excel sheet and determines whether or not the course meets during that time
# Param : "course" - the class section object we want to block out time for
#        "day" - specifies which day we are looking at to compare the class schedule to
#        "currentTime" - the hour and minute that represents the time slot we are comparing the class schedule to
# Returns true if the employee does not have a class during that time
#        false otherwise
def isAvailable(course, day, currentTime):
    days = course["meetingDays"]
    startClass = convert_to_time(course["start"])
    endClass = convert_to_time(course["end"])

    for char in days:
        if char == day:
            if currentTime >= startClass and currentTime < endClass:
                return False

    return True


# Restores an empty grid so that each time this program runs, the grid will be overwritten
# Iterates through each cell and fills it with the empty white color
# No return value or parameters needed
def clearGrid(ws):
    min_row = 3
    min_col = 2
    max_row = 9

    # Iterate through the specified range of rows and columns
    for row in range(min_row, max_row + 1):
        for col in range(min_col, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            fillColor = PatternFill(
                start_color="FFFFFF", end_color="FFFFFF", fill_type="solid"
            )
            cell.fill = fillColor


# Run the program that updates and saves the TimeTable worksheet
def main():
    # Construct the path to the Excel file using os.path.join
    # This dynamically constructs the file path based on the current directory of the script
    excel_file_path = os.path.join(current_dir, "Timetable.xlsx")
    print(f"Excel file path: {excel_file_path}")

    # Load the workbook using the constructed path
    # This replaces the hard-coded path with a dynamically constructed one
    wb = load_workbook(excel_file_path)
    ws = wb.active

    # Check if the file exists at the constructed path
    # This ensures that the script will not try to open a file that doesn't exist, preventing a FileNotFoundError
    if not os.path.exists(excel_file_path):
        print("Error: The file Timetable.xlsx does not exist at the specified path.")
        sys.exit(1)

    # Empty out the grid so that it will be overwritten
    clearGrid(ws)

    # Get the student schedule used to populate the grid
    # TODO: change this parameter once workday api can be accessed
    data1 = getStudentSchedule(0)

    if data1:
        populateGrid(data1, ws)
    else:
        print("Error fetching data from Workday")

    wb.save("Timetable.xlsx")


# Run the main program when gridGenerator is executed
if __name__ == "__main__":
    main()