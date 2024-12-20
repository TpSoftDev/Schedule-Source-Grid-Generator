
import controllers.api_calls.test_site.schedule_source_api
from openpyxl import load_workbook


credentials = {
    "Code" : "isu",
    "Username" : "btowle04",
    "Password" : "6269",
}

wb = load_workbook("../../Timetable template.xlsx")
ws = wb.active
cell = ws.cell(row=1, column=1)
cell.value = "IM IN"